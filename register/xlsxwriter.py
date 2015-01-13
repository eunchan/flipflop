# -*- coding: utf-8 -*-
from __future__ import print_function
from __future__ import unicode_literals
from __future__ import division
from __future__ import absolute_import

from openpyxl import Workbook
from openpyxl.cell import get_column_letter, coordinate_from_string
from openpyxl.style import * # Color, Font, Borders, Align blah blah blah

from .block import Block
from .exception import ProcedureError
from .writer import Writer

class XlsxWriter(Writer):
    BG_HEADER = 'FFCFFFFF'
    BG_HEADER_BIT2 = 'FFCFFFCF'
    BG_HEADER_BIT1 = 'FFFFFFCF'
    C_MIDDLE = 'FFAAAAAA'
    FG_OFFSET = 'FF8B4513'
    FG_REGNAME = 'FF4B0082'
    FG_POR = 'FF8B4513'

    def __init__(self):
        self.wb = Workbook()

    def __init__(self, blocks):
        self.wb = Workbook()
        super(XlsxWriter,self).__init__(blocks)

    def __repr__(self):
        pass

    def save(self, filename):
        if self.ws:
            self.wb.save(filename)
        else:
            raise ProcedureError("build should be run prior to save the "
                    "result into file.")

    def build(self):
        """Builing Excel Worksheet."""

        # Create Worksheet for registers
        self.ws = self.wb.worksheets[0] # optimized_write should not be true
        self.ws.title = 'Registers'

        # Styling
        self.buildStyles()

        # Write Headers
        self.buildHeader(self.ws)

        # While Looping, build registers
        offset = (2,4)
        for block in self.blocks:
            offset = self.buildBlock(self.ws,block,offset)

    def buildStyles(self):
        self.styHeader = Style()
        self.styHeader.fill.fill_type = Fill.FILL_SOLID
        self.styHeader.fill.start_color = Color(XlsxWriter.BG_HEADER)
        self.styHeader.font.size = 8
        self.styHeader.font.bold = True
        self.styHeader.alignment.horizontal = Alignment.HORIZONTAL_CENTER
        self.styHeader.alignment.vertical = Alignment.VERTICAL_CENTER
        self.styHeader.alignment.shrink_to_fit = True
        self._set_borders(self.styHeader, 'LTRB')

        self.styHdrBit1 = Style()
        self.styHdrBit1.fill.fill_type = Fill.FILL_SOLID
        self.styHdrBit1.fill.start_color = Color(self.BG_HEADER_BIT1)
        self.styHdrBit1.font.size = 8
        self.styHdrBit1.font.bold = True
        self.styHdrBit1.alignment.horizontal = Alignment.HORIZONTAL_CENTER
        self.styHdrBit1.alignment.vertical = Alignment.VERTICAL_CENTER
        self.styHdrBit1.alignment.shrink_to_fit = True
        self._set_borders(self.styHdrBit1, 'LTRB')
        
        self.styHdrBit2 = Style()
        self.styHdrBit2.fill.fill_type = Fill.FILL_SOLID
        self.styHdrBit2.fill.start_color = Color(self.BG_HEADER_BIT2)
        self.styHdrBit2.font.size = 8
        self.styHdrBit2.font.bold = True
        self.styHdrBit2.alignment.horizontal = Alignment.HORIZONTAL_CENTER
        self.styHdrBit2.alignment.vertical = Alignment.VERTICAL_CENTER
        self.styHdrBit2.alignment.shrink_to_fit = True
        self._set_borders(self.styHdrBit2, 'LTRB')

        self.styOffset = Style()
        self.styOffset.font.size = 8
        self.styOffset.font.color = Color(self.FG_OFFSET)
        self.styOffset.alignment.horizontal = Alignment.HORIZONTAL_CENTER
        self.styOffset.alignment.vertical = Alignment.VERTICAL_CENTER
        self._set_borders(self.styOffset, 'LTRB')

        self.styRegister = Style()
        self.styRegister.font.size = 8
        self.styRegister.font.bold = True
        self.styRegister.font.color = Color(self.FG_REGNAME)
        self.styRegister.alignment.vertical = Alignment.VERTICAL_CENTER
        self.styRegister.alignment.wrap_text = True
        self.styRegister.alignment.indent = 1
        self._set_borders(self.styRegister, 'LTRB')

        self.styPor1 = Style()
        self.styPor1.font.size = 8
        self.styPor1.font.color = Color(self.FG_POR)
        self.styPor1.alignment.horizontal = Alignment.HORIZONTAL_CENTER
        self.styPor1.alignment.vertical = Alignment.VERTICAL_CENTER
        self._set_borders(self.styPor1,'LTRb')
        self.styPor2 = Style()
        self.styPor2.font.size = 8
        self.styPor2.font.color = Color(self.FG_POR)
        self.styPor2.alignment.horizontal = Alignment.HORIZONTAL_CENTER
        self.styPor2.alignment.vertical = Alignment.VERTICAL_CENTER
        self._set_borders(self.styPor2, 'LtRB')
        
        self.styBit1 = Style()
        self.styBit1.font.size = 8
        self.styBit1.alignment.horizontal = Alignment.HORIZONTAL_CENTER
        self.styBit1.alignment.vertical = Alignment.VERTICAL_CENTER
        self.styBit1.alignment.shrink_to_fit = True
        self._set_borders(self.styBit1, 'LTRb')
        self.styBit2 = Style()
        self.styBit2.font.size = 8
        self.styBit2.alignment.horizontal = Alignment.HORIZONTAL_CENTER
        self.styBit2.alignment.vertical = Alignment.VERTICAL_CENTER
        self.styBit2.alignment.shrink_to_fit = True
        self._set_borders(self.styBit2, 'LtRB')

        # Detail Styling for reserved bit fields
        self.styRsvdLT = Style()
        self.styRsvdT = Style()
        self.styRsvdRT = Style()
        self.styRsvdLB = Style()
        self.styRsvdB = Style()
        self.styRsvdRB = Style()
        self._set_borders(self.styRsvdLT, 'LTb')
        self._set_borders(self.styRsvdT, 'Tb')
        self._set_borders(self.styRsvdRT, 'RTb')
        self._set_borders(self.styRsvdLB, 'LtB')
        self._set_borders(self.styRsvdB, 'tB')
        self._set_borders(self.styRsvdRB, 'RtB')

    def _set_borders(self, style, corners):
        """Set Borders based on corners character.

        Args:
            style (Style)   : Style class
            corners (str)   : Corner string

        corners string is able to have 8 difference chacters (case-sensitive).
        If a character is caplitalized, it mean it should be main line (blakc at
        this moment). If lowercase, it should be gray line colour. And each
        letter indicates the side of borders. Thus, [lLtTrRbB]* are possible
        combinations.

        """

        for c in corners:
            if c == 'l':
                style.borders.left.border_style = Border.BORDER_THIN
                style.borders.left.color = Color(self.C_MIDDLE)
                pass
            elif c == 'L':
                style.borders.left.border_style = Border.BORDER_THIN
                pass
            elif c == 't':
                style.borders.top.border_style = Border.BORDER_THIN
                style.borders.top.color = Color(self.C_MIDDLE)
                pass
            elif c == 'T':
                style.borders.top.border_style = Border.BORDER_THIN
                pass
            elif c == 'r':
                style.borders.right.border_style = Border.BORDER_THIN
                style.borders.right.color = Color(self.C_MIDDLE)
                pass
            elif c == 'R':
                style.borders.right.border_style = Border.BORDER_THIN
                pass
            elif c == 'b':
                style.borders.bottom.border_style = Border.BORDER_THIN
                style.borders.bottom.color = Color(self.C_MIDDLE)
                pass
            elif c == 'B':
                style.borders.bottom.border_style = Border.BORDER_THIN
                pass
            else:
                raise ArgumentError('Unable to handle : %s letter. %s'
                %(c,'_set_borders'))

    def buildHeader(self, ws):
        """Print out header of Compact Register Map"""
        # Offset Address Field
        mergeAndWriteCell(ws, (2,2), (2,3), 'Offset', self.styHeader)
        ws.column_dimensions['B'].width = 6
        
        # Register field 
        mergeAndWriteCell(ws, (3,2), (3,3), 'Register', self.styHeader)
        ws.column_dimensions['C'].width = 12
        
        # Power on Reset Value field
        mergeAndWriteCell(ws, (20,2), (20,3), 'POR', self.styHeader)
        ws.column_dimensions['T'].width = 5
        
        # Bits field
        for x in range(16):
            writeCell(ws, (4+x,2), '%d'%(31-x),self._get_header_bit_style(x))
            writeCell(ws, (4+x,3), '%d'%(15-x),self._get_header_bit_style(x + 4))
            (column,row) = coordinate_from_string(cellStr(4+x,2))
            ws.column_dimensions[column].width = 6 

    def _get_header_bit_style(self, offset):
        if (offset//4)%2 == 0:
            return self.styHdrBit1
        else:
            return self.styHdrBit2

    def buildBlock(self, ws, block, pos):
        """Print registers of one block.

        Args:
            block (Block): a Block information that contains registers
            pos (): Starting cell position
        
        """
        col = pos[0]
        row = pos[1]
        # Print block name
        # The reason why merging cell is to split cells nicely when printing
        mergeAndWriteCell(ws, (col+2,row+1), (col+17,row+1), block.name, DEFAULTS)

        # Print Registers
        row = row + 2
        offset = (col, row)
        for register in block.registers:
            offset = self.buildRegister(ws, register, offset)

        return (col, row) # return next block position

    def buildRegister(self, ws, register, pos):
        """Print a register.

        Args:
            ws (Worksheet): Worksheet of currently active
            register (Register): Register information
            pos (): Position of Starting point

        """
        col = pos[0]
        row = pos[1]

        # Offset
        mergeAndWriteCell(ws, (col, row), (col,row+1), 
                '0x%04x'%register.offset, self.styOffset)

        # Register Name
        mergeAndWriteCell(ws, (col+1,row), (col+1,row+1), 
                register.name, self.styRegister)

        # Reset Value
        writeCell(ws, (col+18, row), '%04xh'%(register.por//0x10000),
                self.styPor1)
        writeCell(ws, (col+18, row+1), '%04xh'%(register.por%0x10000),
                self.styPor2)

        # Print Bits ! Important!
        for bit in register.bits:
            if bit.name.lower() == 'reserved':
                # Below code is decorating reserved bit field.
                # The reason why code seems to be messy is for me to keep the
                # cells seperated (not merging it). So, many cases should be
                # considered.
                #
                # Cases:
                #   1. Upper Left -- not single, not 16th
                #   2. Middle of Upper or Bottom -- not single, not 16th, not
                #      15th, not high or low position
                #   3. Upper Right -- 16th and not single, or low w/ low>=16
                #   4. Single Cell of Upper or Bottom -- high == low or high
                #      and 16th position
                #   5. Bottom Left -- not single, 15th or high w/ high<=15
                #   6. Bottom Right -- not single and low
                for x in range(bit.high-bit.low + 1):
                    p = bit.high - x
                    p_col = col + 2 + 15 - (p%16)
                    p_row = row + 1 - p//16
                    if p == bit.high:
                        if p == bit.low or p == 16:
                            # one reserved cell, need to use Bit1 or Bit2
                            writeCell(ws,
                                    (p_col,p_row),
                                    '',
                                    self.styBit1 if p >= 16 else self.styBit2)
                        else:
                            # LTb or LtB
                            writeCell(ws,
                                    (p_col,p_row),
                                    '',
                                    self.styRsvdLT if p>=16 else
                                    self.styRsvdLB)
                    elif p == bit.low:
                        if p == 15:
                            # Bit2
                            writeCell(ws,
                                    (p_col,p_row),
                                    '',
                                    self.styBit2)
                        else:
                            #RTb or RtB
                            writeCell(ws,
                                    (p_col,p_row),
                                    '',
                                    self.styRsvdRT if p>=16 else
                                    self.styRsvdRB)
                    elif p == 16:
                        writeCell(ws,
                                (p_col,p_row),
                                '',
                                self.styRsvdRT)
                    elif p == 15:
                        writeCell(ws,
                                (p_col,p_row),
                                '',
                                self.styRsvdLB)
                    else:
                        writeCell(ws,
                                (p_col,p_row),
                                '',
                                self.styRsvdT if p>=16 else self.styRsvdB)
                continue #skip reserved bit field
            # call print bit function
            if bit.high >=16 and bit.low <= 15:
                # need to be splitted
                mergeAndWriteCell(ws,
                        (col+2+15-(bit.high%16),row+1-bit.high//16),
                        (col+2+15-(16%16),row+1-16//16),
                        bit.name,
                        self.styBit1)
                mergeAndWriteCell(ws,
                        (col+2+15-(15%16),row+1-15//16),
                        (col+2+15-(bit.low%16),row+1-bit.low//16),
                        bit.name,
                        self.styBit2)

            #elif bit.high == bit.low:
            #    #Single bit handling
            else:
                mergeAndWriteCell(ws,
                        (col+2+15-(bit.high%16),row+1-bit.high//16),
                        (col+2+15-(bit.low%16),row+1-bit.low//16),
                        bit.name,
                        self.styBit1 if bit.high >= 16 else self.styBit2)



        return (pos[0],pos[1] + 2) # return next register position

def writeCell(ws, pos, value, style):
    """Write into a cell.

    Args:
        ws (Worksheet): Worksheet of openpyxl
        pos (): cell position. (col, row)
        value (str): String value of the cell
        style (Style): cell style such as font, alignment, borders

    """
    ws.cell(cellStr(pos[0],pos[1])).value = value
    
    #_style = ws.get_style(cellStr(pos[0],pos[1]))
    #_style = style
    _set_style(ws, pos, style)

def _set_style(ws, pos, style):
    """Set Style for specific cell"""
    ws._styles[cellStr(pos[0],pos[1])] = style

def mergeAndWriteCell(ws, pos1, pos2, value, style):
    """Merge Cells and Write value.

    Args:
        ws (Worksheet): Worksheet of openpyxl
        pos1 (): Start cell position
        pos2 (): End cell position
        value (str): Cell's string value
        style (Style): cell style

    """
    if (pos1[0] != pos2[0]) or (pos1[1] != pos2[1]):
        ws.merge_cells(rangeStr(pos1[0], pos1[1], pos2[0], pos2[1]))

    writeCell(ws, pos1, value, style)

    # styling for each merged cell
    for col in range(pos2[0] - pos1[0] + 1):
        for row in range(pos2[1] - pos1[1] + 1):
            _set_style(ws, (pos1[0]+col,pos1[1]+row), style)

def cellStr(col, row):
    return "%s%s" % (get_column_letter(col), row)

def rangeStr(col1, row1, col2, row2):
    """return range text.

    Returns:
        A1:B3 format
    
    """
    return cellStr(col1,row1) + ':' + cellStr(col2,row2)
